from __future__ import annotations

import csv
import io
import json
import time
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .store import get_evidence, get_refresh_sessions, get_summary

ALLOWED_METRICS = {
    "impressions",
    "clicks",
    "ctr_percent",
    "spend",
    "cpc",
    "cpm",
    "conversions",
    "revenue",
    "cpa",
    "roas",
    "qualification_score",
    "attribution_confidence",
}

DEFAULT_METRICS = [
    "impressions",
    "clicks",
    "ctr_percent",
    "spend",
    "conversions",
    "revenue",
    "cpa",
    "roas",
]

ALLOWED_SECTIONS = {
    "summary",
    "campaigns",
    "line_items",
    "creatives",
    "intelligence",
    "creative_dna",
    "learning_timeline",
}
DEFAULT_SECTIONS = list(ALLOWED_SECTIONS)

SOURCE_LABELS = {
    "manual": "Library / Manual",
    "google_ads": "Google Ads",
    "meta_ads": "Meta Ads",
}

METRIC_LABELS = {
    "impressions": "Impressions",
    "clicks": "Clicks",
    "ctr_percent": "CTR (%)",
    "spend": "Spend",
    "cpc": "CPC",
    "cpm": "CPM",
    "conversions": "Conversions",
    "revenue": "Conversion Value / Revenue",
    "cpa": "CPA",
    "roas": "ROAS",
    "qualification_score": "Qualification Score",
    "attribution_confidence": "Attribution Confidence",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def _clean_list(values: Iterable[str] | None, allowed: set[str], defaults: list[str]) -> list[str]:
    selected = [str(value).strip() for value in (values or []) if str(value).strip() in allowed]
    return selected or list(defaults)


def _timestamp_label(value: Any) -> str:
    try:
        seconds = int(value or 0)
    except (TypeError, ValueError):
        return ""
    if not seconds:
        return ""
    return datetime.fromtimestamp(seconds, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: Any) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _title(value: Any) -> str:
    return str(value or "").replace("_", " ").strip().title()


def _metric_value(item: dict[str, Any], metric: str) -> Any:
    value = item.get(metric)
    if metric in {"ctr_percent", "cpc", "cpm", "spend", "conversions", "revenue", "cpa", "roas"}:
        return round(_to_float(value), 4)
    if metric in {"qualification_score", "attribution_confidence"}:
        return round(_to_float(value), 4)
    if metric in {"impressions", "clicks"}:
        return _to_int(value)
    return value


def _features(item: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    features = item.get("features") or {}
    return features.get("copy") or {}, features.get("image") or {}, features.get("video") or {}


def _raw(item: dict[str, Any]) -> dict[str, Any]:
    value = item.get("raw_metadata") or {}
    return value if isinstance(value, dict) else {}


def _filter_evidence(
    evidence: list[dict[str, Any]],
    *,
    sources: list[str] | None,
    statuses: list[str] | None,
    updated_start: str | None,
    updated_end: str | None,
) -> list[dict[str, Any]]:
    source_set = {value for value in (sources or []) if value}
    status_set = {value for value in (statuses or []) if value}

    start_ts = None
    end_ts = None
    try:
        if updated_start:
            start_ts = int(datetime.fromisoformat(updated_start).replace(tzinfo=timezone.utc).timestamp())
        if updated_end:
            end_ts = int(
                datetime.fromisoformat(updated_end)
                .replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                .timestamp()
            )
    except ValueError as exc:
        raise ValueError("Invalid report date. Use YYYY-MM-DD.") from exc

    rows = []
    for item in evidence:
        if source_set and item.get("source") not in source_set:
            continue
        if status_set and item.get("evidence_status") not in status_set:
            continue
        changed = _to_int(item.get("lastChangedAt") or item.get("updatedAt"))
        if start_ts is not None and changed < start_ts:
            continue
        if end_ts is not None and changed > end_ts:
            continue
        rows.append(item)
    return rows


def _performance_unit_key(item: dict[str, Any]) -> str:
    explicit = str(item.get("performance_unit_id") or "").strip()
    if explicit:
        return explicit
    source = str(item.get("source") or "unknown")
    account = str(item.get("source_account_id") or "")
    campaign = str(item.get("campaign_id") or "")
    raw = _raw(item)
    if source == "google_ads":
        scope = str(item.get("ad_group_id") or raw.get("adId") or raw.get("ad_id") or raw.get("assetGroupId") or "campaign")
    elif source == "meta_ads":
        scope = str(item.get("deployment_id") or raw.get("metaAdId") or raw.get("adId") or item.get("creative_id") or "creative")
    else:
        scope = str(item.get("deployment_id") or item.get("creative_id") or item.get("external_asset_id") or "creative")
    return ":".join([source, account, campaign, scope])


def _aggregate(rows: list[dict[str, Any]], key_fields: list[str], metrics: list[str]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    seen_units: dict[tuple[Any, ...], set[str]] = defaultdict(set)
    for item in rows:
        key = tuple(item.get(field) for field in key_fields)
        if key not in grouped:
            grouped[key] = {field: item.get(field) for field in key_fields}
            grouped[key]["creative_count"] = 0
            grouped[key]["independent_result_count"] = 0
            for metric in metrics:
                grouped[key][metric] = 0.0
        target = grouped[key]
        target["creative_count"] += 1
        unit = _performance_unit_key(item)
        if unit in seen_units[key]:
            continue
        seen_units[key].add(unit)
        target["independent_result_count"] += 1
        for metric in metrics:
            if metric in {"ctr_percent", "cpc", "cpm", "cpa", "roas", "qualification_score", "attribution_confidence"}:
                continue
            target[metric] += _to_float(item.get(metric))

    for key, target in grouped.items():
        impressions = _to_float(target.get("impressions"))
        clicks = _to_float(target.get("clicks"))
        spend = _to_float(target.get("spend"))
        conversions = _to_float(target.get("conversions"))
        revenue = _to_float(target.get("revenue"))
        matching_units = {}
        for item in rows:
            if all(item.get(field) == target.get(field) for field in key_fields):
                matching_units.setdefault(_performance_unit_key(item), item)
        if "ctr_percent" in metrics:
            target["ctr_percent"] = (clicks / impressions * 100) if impressions else 0
        if "cpc" in metrics:
            target["cpc"] = spend / clicks if clicks else 0
        if "cpm" in metrics:
            target["cpm"] = spend / impressions * 1000 if impressions else 0
        if "cpa" in metrics:
            target["cpa"] = spend / conversions if conversions else 0
        if "roas" in metrics:
            target["roas"] = revenue / spend if spend else 0
        for metric in {"qualification_score", "attribution_confidence"} & set(metrics):
            vals = [_to_float(item.get(metric)) for item in matching_units.values()]
            target[metric] = sum(vals) / len(vals) if vals else 0

    return sorted(grouped.values(), key=lambda row: (_to_float(row.get("spend")), _to_int(row.get("impressions"))), reverse=True)

def _creative_rows(evidence: list[dict[str, Any]], metrics: list[str]) -> tuple[list[str], list[list[Any]]]:
    base = [
        "Source", "Account ID", "Campaign ID", "Campaign Name",
        "Line Item ID", "Line Item Name", "Ad ID", "Creative ID",
        "Asset ID", "Kind", "Asset Role", "Platform", "Evidence Status",
        "Headline", "Primary Text", "CTA", "Image URL", "Video ID",
    ]
    headers = base + [METRIC_LABELS[metric] for metric in metrics] + ["First Seen", "Last Changed"]
    rows = []
    for item in evidence:
        copy, _image, video = _features(item)
        raw = _raw(item)
        rows.append([
            SOURCE_LABELS.get(item.get("source"), _title(item.get("source"))),
            item.get("source_account_id"), item.get("campaign_id"), item.get("campaign_name"),
            item.get("ad_group_id"), raw.get("adSetName") or raw.get("ad_group_name") or raw.get("assetGroupName"),
            raw.get("adId") or raw.get("ad_id"), item.get("creative_id"), item.get("external_asset_id"),
            item.get("kind"), item.get("asset_role"), item.get("platform_label"), item.get("evidence_status"),
            copy.get("headline") or raw.get("headline"), copy.get("body") or raw.get("primaryText") or raw.get("body"),
            copy.get("cta") or raw.get("ctaType") or raw.get("cta"), raw.get("imageUrl") or raw.get("previewUrl"),
            raw.get("videoId") or video.get("video_id"),
            *[_metric_value(item, metric) for metric in metrics],
            _timestamp_label(item.get("firstSeenAt")), _timestamp_label(item.get("lastChangedAt")),
        ])
    return headers, rows


def _intelligence_rows(evidence: list[dict[str, Any]], metrics: list[str]) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "Evidence ID", "Source", "Campaign ID", "Line Item ID", "Creative ID", "Asset ID",
        "Evidence Status", "Visual Style", "Composition", "Background", "Imagery Type",
        "Emotional Tone", "Dominant Colors", "Headline Opener", "CTA Opener",
        "Headline Length", "Product Prominence (%)",
    ] + [METRIC_LABELS[metric] for metric in metrics]
    rows = []
    for item in evidence:
        copy, image, _video = _features(item)
        rows.append([
            item.get("id"), SOURCE_LABELS.get(item.get("source"), _title(item.get("source"))),
            item.get("campaign_id"), item.get("ad_group_id"), item.get("creative_id"), item.get("external_asset_id"),
            item.get("evidence_status"), image.get("visual_style"), image.get("composition"),
            image.get("background_type"), image.get("lifestyle_vs_studio"), image.get("emotional_tone"),
            ", ".join(str(value) for value in (image.get("dominant_colors") or [])),
            copy.get("first_headline_word"), copy.get("first_cta_word"), copy.get("headline_length"),
            image.get("product_prominence_percent"),
            *[_metric_value(item, metric) for metric in metrics],
        ])
    return headers, rows


def _dna_rows(summary: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    profile = summary.get("generationProfile") or {}
    rows: list[list[Any]] = []
    mappings = [
        ("Top Colors", "top_colors"), ("Visual Styles", "top_visual_styles"),
        ("Compositions", "top_compositions"), ("Backgrounds", "top_backgrounds"),
        ("Imagery Types", "top_imagery_types"), ("Emotional Tones", "top_emotional_tones"),
        ("CTA Openers", "top_cta_openers"), ("Headline Openers", "top_headline_openers"),
        ("Asset Roles", "top_asset_roles"),
    ]
    for category, key in mappings:
        for rank, item in enumerate(profile.get(key) or [], start=1):
            rows.append([category, rank, item.get("value"), item.get("count"), item.get("share")])
    rows.extend([
        ["Format Metric", 1, "Average winning headline length", profile.get("average_winning_headline_length"), None],
        ["Format Metric", 2, "Average winning product prominence (%)", profile.get("average_winning_product_prominence_percent"), None],
    ])
    return ["Category", "Rank", "Value", "Weighted Count", "Share"], rows


def _timeline_rows(sessions: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "Refresh ID", "Status", "Started", "Finished", "Duration (seconds)",
        "Evidence Before", "Evidence After", "Confidence Before", "Confidence After",
        "New", "Updated", "Unchanged", "Failures", "Recommendation",
    ]
    rows = []
    for session in sessions:
        before = session.get("before") or {}
        after = session.get("after") or {}
        changes = session.get("learningChanges") or {}
        sources = session.get("sources") or {}
        rows.append([
            session.get("id"), session.get("status"), _timestamp_label(session.get("startedAt")),
            _timestamp_label(session.get("finishedAt")), session.get("durationSeconds"),
            before.get("evidenceCount"), after.get("evidenceCount"), before.get("confidence"), after.get("confidence"),
            changes.get("added", sum(_to_int(v.get("added")) for v in sources.values() if isinstance(v, dict))),
            changes.get("updated", sum(_to_int(v.get("updated")) for v in sources.values() if isinstance(v, dict))),
            changes.get("unchanged", sum(_to_int(v.get("unchanged")) for v in sources.values() if isinstance(v, dict))),
            changes.get("failureCount", sum(len(v.get("failures") or []) for v in sources.values() if isinstance(v, dict))),
            changes.get("recommendation"),
        ])
    return headers, rows


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max(1, len(headers)))}{max(1, len(rows) + 1)}"
    for row in rows:
        ws.append(row)
    for index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for cell in ws[get_column_letter(index)]:
            max_length = max(max_length, min(len(str(cell.value or "")), 80))
        ws.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 48)
    ws.sheet_view.showGridLines = False


def _summary_sheet(ws, summary: dict[str, Any], evidence: list[dict[str, Any]], metrics: list[str], filters: dict[str, Any]) -> None:
    ws["A1"] = "ADGen Performance Report"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A3"] = "Generated"
    ws["B3"] = _timestamp_label(int(time.time()))
    ws["A4"] = "Scope"
    ws["B4"] = "Accumulated Performance Intelligence evidence"
    ws["A5"] = "Important"
    ws["B5"] = "Campaign and line-item totals are creative-attributed aggregates and may not equal provider billing totals."
    ws["A7"] = "Filters"
    ws["A7"].fill = SUBHEADER_FILL
    ws["A7"].font = BOLD_FONT
    ws["A8"] = "Sources"
    ws["B8"] = ", ".join(filters.get("sources") or ["All"])
    ws["A9"] = "Evidence statuses"
    ws["B9"] = ", ".join(filters.get("statuses") or ["All"])
    ws["A10"] = "Evidence updated start"
    ws["B10"] = filters.get("updated_start") or "Any"
    ws["A11"] = "Evidence updated end"
    ws["B11"] = filters.get("updated_end") or "Any"
    ws["A13"] = "Learning Summary"
    ws["A13"].fill = SUBHEADER_FILL
    ws["A13"].font = BOLD_FONT
    summary_rows = [
        ("Evidence in export", len(evidence)), ("Overall retained evidence", summary.get("evidenceCount", 0)),
        ("Qualified evidence", summary.get("qualifiedCount", 0)), ("Positive signals", summary.get("positiveCount", 0)),
        ("Underperformers", summary.get("underperformerCount", 0)), ("Confidence", summary.get("confidence", 0)),
        ("Average positive CTR (%)", summary.get("averagePositiveCtrPercent")),
        ("Average positive ROAS", summary.get("averagePositiveRoas")),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=14):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
    ws["D13"] = "Included Metrics"
    ws["D13"].fill = SUBHEADER_FILL
    ws["D13"].font = BOLD_FONT
    for row_index, metric in enumerate(metrics, start=14):
        ws.cell(row=row_index, column=4, value=METRIC_LABELS[metric])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["D"].width = 30
    ws.sheet_view.showGridLines = False


def build_report_data(
    uid: str,
    *,
    metrics: list[str] | None = None,
    sections: list[str] | None = None,
    sources: list[str] | None = None,
    statuses: list[str] | None = None,
    updated_start: str | None = None,
    updated_end: str | None = None,
) -> dict[str, Any]:
    selected_metrics = _clean_list(metrics, ALLOWED_METRICS, DEFAULT_METRICS)
    selected_sections = _clean_list(sections, ALLOWED_SECTIONS, DEFAULT_SECTIONS)
    all_evidence = get_evidence(uid, limit=5000)
    evidence = _filter_evidence(
        all_evidence,
        sources=sources,
        statuses=statuses,
        updated_start=updated_start,
        updated_end=updated_end,
    )
    summary = get_summary(uid)
    sessions = get_refresh_sessions(uid, limit=100)

    campaign_rows = _aggregate(
        evidence,
        ["source", "source_account_id", "campaign_id", "campaign_name"],
        selected_metrics,
    )
    line_item_rows = _aggregate(
        evidence,
        ["source", "source_account_id", "campaign_id", "campaign_name", "ad_group_id"],
        selected_metrics,
    )
    return {
        "metrics": selected_metrics,
        "sections": selected_sections,
        "filters": {"sources": sources or [], "statuses": statuses or [], "updated_start": updated_start, "updated_end": updated_end},
        "summary": summary,
        "evidence": evidence,
        "sessions": sessions,
        "campaigns": campaign_rows,
        "line_items": line_item_rows,
    }


def build_excel_report(uid: str, **kwargs: Any) -> tuple[bytes, str]:
    data = build_report_data(uid, **kwargs)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "ADGen MCM"
    workbook.properties.title = "ADGen Performance Report"
    workbook.properties.description = "Performance Intelligence reporting export"

    sections = set(data["sections"])
    metrics = data["metrics"]
    if "summary" in sections:
        _summary_sheet(workbook.create_sheet("Summary"), data["summary"], data["evidence"], metrics, data["filters"])

    if "campaigns" in sections:
        headers = ["Source", "Account ID", "Campaign ID", "Campaign Name", "Creative Count"] + [METRIC_LABELS[m] for m in metrics]
        rows = [[SOURCE_LABELS.get(r.get("source"), _title(r.get("source"))), r.get("source_account_id"), r.get("campaign_id"), r.get("campaign_name"), r.get("creative_count"), *[_metric_value(r, m) for m in metrics]] for r in data["campaigns"]]
        _write_sheet(workbook.create_sheet("Campaigns"), headers, rows)

    if "line_items" in sections:
        headers = ["Source", "Account ID", "Campaign ID", "Campaign Name", "Line Item ID", "Creative Count"] + [METRIC_LABELS[m] for m in metrics]
        rows = [[SOURCE_LABELS.get(r.get("source"), _title(r.get("source"))), r.get("source_account_id"), r.get("campaign_id"), r.get("campaign_name"), r.get("ad_group_id"), r.get("creative_count"), *[_metric_value(r, m) for m in metrics]] for r in data["line_items"]]
        _write_sheet(workbook.create_sheet("Line Items"), headers, rows)

    if "creatives" in sections:
        headers, rows = _creative_rows(data["evidence"], metrics)
        _write_sheet(workbook.create_sheet("Creatives"), headers, rows)

    if "intelligence" in sections:
        headers, rows = _intelligence_rows(data["evidence"], metrics)
        _write_sheet(workbook.create_sheet("Intelligence"), headers, rows)

    if "creative_dna" in sections:
        headers, rows = _dna_rows(data["summary"])
        _write_sheet(workbook.create_sheet("Creative DNA"), headers, rows)

    if "learning_timeline" in sections:
        headers, rows = _timeline_rows(data["sessions"])
        _write_sheet(workbook.create_sheet("Learning Timeline"), headers, rows)

    if not workbook.sheetnames:
        _summary_sheet(workbook.create_sheet("Summary"), data["summary"], data["evidence"], metrics, data["filters"])

    stream = io.BytesIO()
    workbook.save(stream)
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    return stream.getvalue(), f"ADGen_Performance_Report_{timestamp}.xlsx"


def build_csv_zip(uid: str, **kwargs: Any) -> tuple[bytes, str]:
    data = build_report_data(uid, **kwargs)
    files: dict[str, tuple[list[str], list[list[Any]]]] = {}
    metrics = data["metrics"]
    sections = set(data["sections"])
    if "campaigns" in sections:
        headers = ["Source", "Account ID", "Campaign ID", "Campaign Name", "Creative Count"] + [METRIC_LABELS[m] for m in metrics]
        rows = [[SOURCE_LABELS.get(r.get("source"), _title(r.get("source"))), r.get("source_account_id"), r.get("campaign_id"), r.get("campaign_name"), r.get("creative_count"), *[_metric_value(r, m) for m in metrics]] for r in data["campaigns"]]
        files["campaigns.csv"] = (headers, rows)
    if "line_items" in sections:
        headers = ["Source", "Account ID", "Campaign ID", "Campaign Name", "Line Item ID", "Creative Count"] + [METRIC_LABELS[m] for m in metrics]
        rows = [[SOURCE_LABELS.get(r.get("source"), _title(r.get("source"))), r.get("source_account_id"), r.get("campaign_id"), r.get("campaign_name"), r.get("ad_group_id"), r.get("creative_count"), *[_metric_value(r, m) for m in metrics]] for r in data["line_items"]]
        files["line_items.csv"] = (headers, rows)
    if "creatives" in sections:
        files["creatives.csv"] = _creative_rows(data["evidence"], metrics)
    if "intelligence" in sections:
        files["intelligence.csv"] = _intelligence_rows(data["evidence"], metrics)
    if "creative_dna" in sections:
        files["creative_dna.csv"] = _dna_rows(data["summary"])
    if "learning_timeline" in sections:
        files["learning_timeline.csv"] = _timeline_rows(data["sessions"])

    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("report_metadata.json", json.dumps({"generatedAt": int(time.time()), **data["filters"], "metrics": metrics, "sections": data["sections"]}, indent=2))
        for filename, (headers, rows) in files.items():
            text = io.StringIO()
            writer = csv.writer(text)
            writer.writerow(headers)
            writer.writerows(rows)
            archive.writestr(filename, text.getvalue())
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    return stream.getvalue(), f"ADGen_Performance_Report_{timestamp}.zip"
