from __future__ import annotations
import io
from datetime import datetime, timezone
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from .metrics import METRICS

def _style(ws):
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for c in ws[1]: c.fill=PatternFill("solid",fgColor="312E81"); c.font=Font(color="FFFFFF",bold=True); c.alignment=Alignment(vertical="center")
    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width=min(42,max(11,max(len(str(c.value or "")) for c in col)+2))

def build_workbook(report:dict[str,Any]) -> io.BytesIO:
    wb=Workbook(); ws=wb.active; ws.title="Executive Summary"
    ws.append(["ADGen Report","Value"]); ws.append(["Generated",datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")]); ws.append(["Reporting period",report["dateRange"]["label"]]); ws.append(["Sources",", ".join(report["providers"])]); ws.append(["Split by"," → ".join(report["splits"])])
    for m,v in report["totals"].items(): ws.append([METRICS.get(m,m),v])
    for notice in report.get("notices",[]): ws.append(["Data note",notice])
    _style(ws)
    ws=wb.create_sheet("Performance"); columns=report["columns"]; ws.append([METRICS.get(c,c.replace("_"," ").title()) for c in columns])
    for row in report["rows"]: ws.append([row.get(c,0) for c in columns])
    _style(ws)
    stream=io.BytesIO(); wb.save(stream); stream.seek(0); return stream
